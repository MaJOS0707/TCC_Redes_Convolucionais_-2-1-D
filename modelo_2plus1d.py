# -*- coding: utf-8 -*-
"""
Treinamento Sequencial - Rede Neural Convolucional (2+1)D
Classificacao de Videos: IA vs Real

Ambiente: Google Colab Pro+ (GPU T4)
"""

# ============================================================
# Etapa 1 - Conexao com Drive e GPU
# ============================================================
# !pip install einops openpyxl -q
# from google.colab import drive
# drive.mount('/content/drive')
import tensorflow as tf
gpus = tf.config.list_physical_devices('GPU')
print(f"GPU: {gpus[0]}" if gpus else "SEM GPU!")

# ============================================================
# Etapa 2 - Imports e Configuracoes
# ============================================================
import pathlib, random, time, os
import cv2, einops
import numpy as np
import matplotlib.pyplot as plt
import keras
from keras import layers
import tensorflow as tf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# Diretorios (ajuste conforme seu ambiente)
DATASET_DIR    = pathlib.Path('/content/drive/MyDrive/TCC_AIGVDBench/dataset_2025_split_min76')
RESULTADOS_DIR = pathlib.Path('/content/drive/MyDrive/TCC_AIGVDBench/resultados')

# Amostras por split e por classe (None = usa todos os videos)
AMOSTRAS = {
    'train': {'ai': None, 'real': None},
    'val':   {'ai': None, 'real': None},
    'test':  {'ai': None, 'real': None},
}

RANDOM_SEED = 122

# Modo de redimensionamento dos frames:
# "stretch"     -> redimensiona direto, sem padding
# "center_crop" -> corta quadrado central e redimensiona
# "pad"         -> preserva proporcao com padding preto
RESIZE_MODE = "stretch"

# True  -> grayscale (1 canal)
# False -> RGB (3 canais)
PRETO_E_BRANCO = False

# Range de frames por video (loop de treino)
FRAME_MIN = 1
FRAME_MAX = 15
DURACAO_MAX_SEC = None   # None = usa o video inteiro

# Hiperparametros
FRAME_STEP    = 4
HEIGHT        = 112
WIDTH         = 112
BATCH_SIZE    = 20
EPOCHS        = 25
LEARNING_RATE = 0.01

VIDEO_EXTS = ["*.avi", "*.mp4", "*.mov", "*.mkv"]
RUN_TAG = datetime.now().strftime("%Y%m%d_%H%M")

# ============================================================
# Etapa 3 - Funcoes de Pre-processamento
# ============================================================

def resize_frame(frame, output_size, mode):
    """Redimensiona um frame de acordo com o modo selecionado."""
    if mode == "stretch":
        return tf.image.resize(frame, output_size)
    elif mode == "center_crop":
        shape = tf.shape(frame)
        h, w = shape[0], shape[1]
        side = tf.minimum(h, w)
        frame = tf.image.resize_with_crop_or_pad(frame, side, side)
        return tf.image.resize(frame, output_size)
    elif mode == "pad":
        return tf.image.resize_with_pad(frame, *output_size)
    else:
        raise ValueError(f"RESIZE_MODE desconhecido: {mode}")


def format_frames(frame, output_size, resize_mode=None):
    """Normaliza para [0, 1] e redimensiona conforme RESIZE_MODE."""
    if resize_mode is None:
        resize_mode = RESIZE_MODE
    frame = tf.image.convert_image_dtype(frame, tf.float32)
    frame = resize_frame(frame, output_size, resize_mode)
    return frame


def frames_from_video_file(video_path, n_frames, output_size=(224, 224),
                           frame_step=15, deterministic=False):
    """
    Le n_frames frames de um video, espacados por frame_step.
    Se o video for menor que o necessario, preenche com frames vazios.
    """
    result = []
    src = cv2.VideoCapture(str(video_path))
    video_length = src.get(cv2.CAP_PROP_FRAME_COUNT)
    fps = src.get(cv2.CAP_PROP_FPS)

    if DURACAO_MAX_SEC is not None and fps > 0:
        max_frames = int(fps * DURACAO_MAX_SEC)
        video_length = min(video_length, max_frames)

    need_length = 1 + (n_frames - 1) * frame_step

    if need_length > video_length:
        start = 0
    elif deterministic:
        start = max(0, int((video_length - need_length) // 2))
    else:
        start = random.randint(0, int(video_length - need_length + 1))

    src.set(cv2.CAP_PROP_POS_FRAMES, start)
    ret, frame = src.read()

    if not ret:
        src.release()
        n_channels = 1 if PRETO_E_BRANCO else 3
        return np.zeros((n_frames, *output_size, n_channels), dtype=np.float32)

    result.append(format_frames(frame, output_size))
    for _ in range(n_frames - 1):
        for _ in range(frame_step):
            ret, frame = src.read()
        if ret:
            result.append(format_frames(frame, output_size))
        else:
            result.append(np.zeros_like(result[0]))
    src.release()

    result = np.array(result)[..., [2, 1, 0]]  # BGR -> RGB

    if PRETO_E_BRANCO:
        result = tf.image.rgb_to_grayscale(result).numpy()

    return result


class FrameGenerator:
    """Gerador de frames para uso com tf.data.Dataset.from_generator."""

    def __init__(self, path, n_frames, training=False):
        self.path = path
        self.n_frames = n_frames
        self.training = training
        self.deterministic = not training
        self.split_name = path.name
        self.class_names = sorted(
            set(p.name for p in self.path.iterdir() if p.is_dir())
        )
        self.class_ids_for_name = {
            name: idx for idx, name in enumerate(self.class_names)
        }

    def get_files_and_class_names(self):
        video_paths = []
        for ext in VIDEO_EXTS:
            video_paths.extend(self.path.glob(f"*/{ext}"))
        classes = [p.parent.name for p in video_paths]
        return video_paths, classes

    def _build_pairs(self):
        """Aplica amostragem por classe com seed fixa."""
        video_paths, classes = self.get_files_and_class_names()
        pairs = list(zip(video_paths, classes))
        rng = random.Random(RANDOM_SEED)
        limites = AMOSTRAS.get(self.split_name, {})
        por_classe = {}
        for path, cls in pairs:
            por_classe.setdefault(cls, []).append((path, cls))
        pairs = []
        for cls, items in sorted(por_classe.items()):
            rng.shuffle(items)
            limite = limites.get(cls, None)
            if limite is not None:
                pairs.extend(items[:limite])
            else:
                pairs.extend(items)
        return pairs

    def first_video_path(self):
        pairs = self._build_pairs()
        return pairs[0][0] if pairs else None

    def __call__(self):
        pairs = self._build_pairs()
        if self.training:
            random.shuffle(pairs)
        for path, name in pairs:
            video_frames = frames_from_video_file(
                path, self.n_frames,
                output_size=(HEIGHT, WIDTH),
                frame_step=FRAME_STEP,
                deterministic=self.deterministic,
            )
            label = self.class_ids_for_name[name]
            yield video_frames, label


# ============================================================
# Etapa 4 - Arquitetura do Modelo (2+1)D
# ============================================================

@keras.saving.register_keras_serializable()
class Conv2Plus1D(keras.layers.Layer):
    """Convolucao fatorizada: espacial 2D seguida de temporal 1D."""

    def __init__(self, filters, kernel_size, padding, **kwargs):
        super().__init__(**kwargs)
        self.filters = filters
        self.kernel_size = kernel_size
        self.padding = padding
        self.seq = keras.Sequential([
            layers.Conv3D(filters=filters,
                          kernel_size=(1, kernel_size[1], kernel_size[2]),
                          padding=padding),
            layers.Conv3D(filters=filters,
                          kernel_size=(kernel_size[0], 1, 1),
                          padding=padding),
        ])

    def call(self, x):
        return self.seq(x)

    def get_config(self):
        config = super().get_config()
        config.update({"filters": self.filters,
                        "kernel_size": self.kernel_size,
                        "padding": self.padding})
        return config


@keras.saving.register_keras_serializable()
class ResidualMain(keras.layers.Layer):
    """Bloco residual principal com duas camadas Conv2Plus1D."""

    def __init__(self, filters, kernel_size, **kwargs):
        super().__init__(**kwargs)
        self.filters = filters
        self.kernel_size = kernel_size
        self.seq = keras.Sequential([
            Conv2Plus1D(filters=filters, kernel_size=kernel_size,
                        padding="same"),
            layers.LayerNormalization(),
            layers.ReLU(),
            Conv2Plus1D(filters=filters, kernel_size=kernel_size,
                        padding="same"),
            layers.LayerNormalization(),
        ])

    def call(self, x):
        return self.seq(x)

    def get_config(self):
        config = super().get_config()
        config.update({"filters": self.filters,
                        "kernel_size": self.kernel_size})
        return config


@keras.saving.register_keras_serializable()
class Project(keras.layers.Layer):
    """Projecao linear para ajuste de dimensoes na conexao residual."""

    def __init__(self, units, **kwargs):
        super().__init__(**kwargs)
        self.units = units
        self.seq = keras.Sequential([
            layers.Dense(units),
            layers.LayerNormalization(),
        ])

    def call(self, x):
        return self.seq(x)

    def get_config(self):
        config = super().get_config()
        config.update({"units": self.units})
        return config


@keras.saving.register_keras_serializable()
class ResizeVideo(keras.layers.Layer):
    """Redimensiona espacialmente cada frame de um tensor de video."""

    def __init__(self, height, width, **kwargs):
        super().__init__(**kwargs)
        self.height = height
        self.width = width
        self.resizing_layer = layers.Resizing(self.height, self.width)

    def call(self, video):
        old_shape = einops.parse_shape(video, "b t h w c")
        images = einops.rearrange(video, "b t h w c -> (b t) h w c")
        images = self.resizing_layer(images)
        return einops.rearrange(images, "(b t) h w c -> b t h w c",
                                t=old_shape["t"])

    def get_config(self):
        config = super().get_config()
        config.update({"height": self.height, "width": self.width})
        return config


def add_residual_block(input_tensor, filters, kernel_size):
    """Adiciona um bloco residual ao grafo do modelo."""
    out = ResidualMain(filters, kernel_size)(input_tensor)
    res = input_tensor
    if out.shape[-1] != input_tensor.shape[-1]:
        res = Project(out.shape[-1])(res)
    return layers.add([res, out])


def build_model(num_classes, n_frames, n_channels=3):
    """Constroi o modelo (2+1)D com conexoes residuais."""
    input_shape = (n_frames, HEIGHT, WIDTH, n_channels)
    inp = layers.Input(shape=input_shape)
    x = Conv2Plus1D(filters=16, kernel_size=(3, 7, 7), padding="same")(inp)
    x = layers.BatchNormalization()(x)
    x = layers.ReLU()(x)
    x = ResizeVideo(HEIGHT // 2,  WIDTH // 2)(x)
    x = add_residual_block(x, 16,  (3, 3, 3))
    x = ResizeVideo(HEIGHT // 4,  WIDTH // 4)(x)
    x = add_residual_block(x, 32,  (3, 3, 3))
    x = ResizeVideo(HEIGHT // 8,  WIDTH // 8)(x)
    x = add_residual_block(x, 64,  (3, 3, 3))
    x = ResizeVideo(HEIGHT // 16, WIDTH // 16)(x)
    x = add_residual_block(x, 128, (3, 3, 3))
    x = layers.GlobalAveragePooling3D()(x)
    x = layers.Flatten()(x)
    x = layers.Dense(num_classes)(x)
    return keras.Model(inp, x)


# ============================================================
# Etapa 5 - Criacao dos Datasets
# ============================================================

def create_datasets(n_frames):
    n_channels = 1 if PRETO_E_BRANCO else 3
    output_signature = (
        tf.TensorSpec(shape=(None, None, None, n_channels), dtype=tf.float32),
        tf.TensorSpec(shape=(), dtype=tf.int16),
    )
    fg_train = FrameGenerator(DATASET_DIR / "train", n_frames, training=True)
    fg_val   = FrameGenerator(DATASET_DIR / "val",   n_frames)
    fg_test  = FrameGenerator(DATASET_DIR / "test",  n_frames)

    train_ds = (tf.data.Dataset.from_generator(fg_train,
                    output_signature=output_signature)
                .batch(BATCH_SIZE).prefetch(tf.data.AUTOTUNE))
    val_ds   = (tf.data.Dataset.from_generator(fg_val,
                    output_signature=output_signature)
                .batch(BATCH_SIZE).prefetch(tf.data.AUTOTUNE))
    test_ds  = (tf.data.Dataset.from_generator(fg_test,
                    output_signature=output_signature)
                .batch(BATCH_SIZE).prefetch(tf.data.AUTOTUNE))

    return train_ds, val_ds, test_ds, fg_train.class_names, fg_train, fg_test


def salvar_configuracoes_txt(out_path, n_frames, n_channels):
    """Salva um TXT com as configuracoes utilizadas no modelo."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    linhas = [
        "=" * 60,
        " CONFIGURACOES DO MODELO",
        "=" * 60,
        f"DATASET_DIR     = {DATASET_DIR}",
        f"RESIZE_MODE     = {RESIZE_MODE}",
        f"PRETO_E_BRANCO  = {PRETO_E_BRANCO}",
        f"HEIGHT x WIDTH  = {HEIGHT} x {WIDTH}",
        f"N_CHANNELS      = {n_channels}",
        f"N_FRAMES        = {n_frames}",
        f"FRAME_STEP      = {FRAME_STEP}",
        f"DURACAO_MAX_SEC = {DURACAO_MAX_SEC}",
        f"RANDOM_SEED     = {RANDOM_SEED}",
        f"BATCH_SIZE      = {BATCH_SIZE}",
        f"EPOCHS (max)    = {EPOCHS}",
        f"LEARNING_RATE   = {LEARNING_RATE}",
        "=" * 60,
    ]
    out_path.write_text("\n".join(linhas), encoding="utf-8")


# ============================================================
# Etapa 6 - Criacao do Excel de Registro
# ============================================================

tipo = "pb" if PRETO_E_BRANCO else "col"
excel_path = (RESULTADOS_DIR /
    f"registro_{tipo}_{RESIZE_MODE}_F{FRAME_MIN}_a_F{FRAME_MAX}_{RUN_TAG}.xlsx")
RESULTADOS_DIR.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Resultados"

header_font  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
header_fill  = PatternFill("solid", fgColor="4472C4")
header_align = Alignment(horizontal="center", vertical="center")
thin_border  = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

headers = ["Modelo", "N_Frames", "Canais", "Resize_Mode",
           "Epocas Executadas",
           "Acc Treino", "Loss Treino",
           "Acc Validacao", "Loss Validacao",
           "Acc Teste", "Loss Teste",
           "Tempo (min)", "Early Stop"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font  = header_font
    cell.fill  = header_fill
    cell.alignment = header_align
    cell.border    = thin_border

wb.save(excel_path)


# ============================================================
# Etapa 7 - Loop de Treino
# ============================================================

data_font  = Font(name="Arial", size=10)
data_align = Alignment(horizontal="center")

for n_frames in range(FRAME_MIN, FRAME_MAX + 1):
    nome_modelo = f"modelo_{tipo}_{RESIZE_MODE}_F{n_frames}_{RUN_TAG}"
    modelo_dir  = RESULTADOS_DIR / nome_modelo
    modelo_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*65}")
    print(f"  TREINANDO: {nome_modelo}  ({n_frames} frames)")
    print(f"  [{n_frames - FRAME_MIN + 1}/{FRAME_MAX - FRAME_MIN + 1}]")
    print(f"{'='*65}\n")

    # --- Datasets ---
    train_ds, val_ds, test_ds, class_names, fg_train, fg_test = \
        create_datasets(n_frames)
    num_classes = len(class_names)
    n_channels  = 1 if PRETO_E_BRANCO else 3
    print(f"Classes: {', '.join(class_names)}")

    salvar_configuracoes_txt(
        modelo_dir / f"{nome_modelo}_config.txt", n_frames, n_channels
    )

    # --- Construir e compilar modelo ---
    model = build_model(num_classes, n_frames, n_channels=n_channels)
    model.compile(
        loss=keras.losses.SparseCategoricalCrossentropy(from_logits=True),
        optimizer=keras.optimizers.Adam(learning_rate=LEARNING_RATE),
        metrics=["accuracy"],
    )

    # --- Treinar ---
    start_time = time.time()
    history = model.fit(
        x=train_ds, epochs=EPOCHS, validation_data=val_ds, verbose=1
    )
    elapsed_min = (time.time() - start_time) / 60

    # --- Metricas finais ---
    epocas_rodadas = len(history.history['loss'])
    train_acc  = history.history['accuracy'][-1]
    train_loss = history.history['loss'][-1]
    val_acc    = history.history['val_accuracy'][-1]
    val_loss   = history.history['val_loss'][-1]
    early_stopped = "Sim" if epocas_rodadas < EPOCHS else "Nao"

    test_loss, test_acc = model.evaluate(test_ds, verbose=0)

    # --- Salvar modelo ---
    caminho_modelo = modelo_dir / f"{nome_modelo}.keras"
    model.save(caminho_modelo)

    # --- Grafico de treino ---
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8))
    fig.suptitle(f"Resultados - {nome_modelo}", fontsize=16)
    ax1.set_title("Loss")
    ax1.plot(history.history["loss"],     label="Treino",    linewidth=2)
    ax1.plot(history.history["val_loss"], label="Validacao", linewidth=2)
    ax1.set_ylabel("Loss"); ax1.set_xlabel("Epoca")
    ax1.legend(); ax1.grid(True, alpha=0.3)
    ax2.set_title("Acuracia")
    ax2.plot(history.history["accuracy"],     label="Treino",    linewidth=2)
    ax2.plot(history.history["val_accuracy"], label="Validacao", linewidth=2)
    ax2.set_ylabel("Acuracia"); ax2.set_xlabel("Epoca")
    ax2.set_ylim([0, 1])
    ax2.legend(); ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(modelo_dir / f"{nome_modelo}_treino.png",
                dpi=150, bbox_inches="tight")
    plt.close()

    # --- Registrar no Excel ---
    row = n_frames - FRAME_MIN + 2
    valores = [
        nome_modelo, n_frames, n_channels, RESIZE_MODE, epocas_rodadas,
        round(train_acc * 100, 5), round(train_loss, 5),
        round(val_acc  * 100, 5), round(val_loss,  5),
        round(test_acc * 100, 5), round(test_loss, 5),
        round(elapsed_min, 5), early_stopped
    ]
    for col, val in enumerate(valores, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = data_font
        cell.alignment = data_align
        cell.border    = thin_border

    wb.save(excel_path)

    # --- Liberar memoria ---
    del model, train_ds, val_ds, test_ds, history, fg_train, fg_test
    keras.backend.clear_session()
    tf.keras.backend.clear_session()

    print(f"\n  Concluido: {nome_modelo}")
    print(f"  Tempo: {elapsed_min:.1f} min | Epocas: {epocas_rodadas}/{EPOCHS}")
    print(f"  Treino:  {train_acc*100:.2f}% | "
          f"Val: {val_acc*100:.2f}% | "
          f"Teste: {test_acc*100:.2f}%")

print(f"\n{'#'*65}")
print(f"#  TODOS OS MODELOS CONCLUIDOS!")
print(f"#  Range:       F{FRAME_MIN} a F{FRAME_MAX}")
print(f"#  Resize mode: {RESIZE_MODE}")
print(f"#  Canais:      {'grayscale' if PRETO_E_BRANCO else 'RGB'}")
print(f"#  Excel:       {excel_path}")
print(f"{'#'*65}")
